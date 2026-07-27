from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from app.extractors.base import BaseExtractor, ExtractedDocument, Heading, Paragraph, Table


class ExcelExtractor(BaseExtractor):
    def extract(self, file_bytes: bytes, filename: str, content_type: str) -> ExtractedDocument:
        doc_hash = self._hash_bytes(file_bytes)
        wb = load_workbook(filename=file_bytes, read_only=False, data_only=True)
        tables: list[Table] = []
        paragraphs: list[Paragraph] = []
        headings: list[Heading] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [self._cell_str(c) for c in row]
                rows.append(cells)

            clean_rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not clean_rows:
                continue

            headers = clean_rows[0]
            data = clean_rows[1:]
            for h in headers:
                if h.strip():
                    headings.append(Heading(level=2, text=h))

            for row in data:
                parts = []
                for i, header in enumerate(headers):
                    if i < len(row) and row[i].strip():
                        parts.append(f"{header}: {row[i]}")
                if parts:
                    paragraphs.append(Paragraph(text=" | ".join(parts)))

            tables.append(
                Table(
                    headers=headers,
                    rows=data,
                    caption=sheet_name,
                )
            )

        title = filename.rsplit(".", 1)[0] if "." in filename else filename

        return ExtractedDocument(
            id=None,
            title=title,
            headings=headings,
            paragraphs=paragraphs,
            tables=tables,
            source=filename,
            hash=doc_hash,
        )

    @staticmethod
    def _cell_str(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
